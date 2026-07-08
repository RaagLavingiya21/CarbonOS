"""Render a neutral `Disclosure` to PDF (fpdf2) or XLSX (openpyxl).

Pure; fpdf/openpyxl are imported lazily. Every dynamic string flows through the
single `_pdf_safe` Latin-1 choke point, so any regime's Unicode facility names
are safe without each mapper re-solving it.
"""

from __future__ import annotations

from io import BytesIO

from s1_reporting.disclosures.ir import Disclosure, Section


def _pdf_safe(value: object) -> str:
    """fpdf2 core fonts are Latin-1 only; make arbitrary user text safe."""
    return str(value).encode("latin-1", "replace").decode("latin-1")


def _fmt(cell: object) -> str:
    """Numbers get thousands separators + 4dp; everything else is stringified."""
    if isinstance(cell, float):
        return f"{cell:,.4f}"
    return str(cell)


def render_pdf(doc: Disclosure) -> bytes:
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 9, _pdf_safe(doc.doc_title), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(90, 90, 90)
    pdf.cell(0, 6, _pdf_safe(doc.subtitle), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)

    for section in doc.sections:
        _render_pdf_section(pdf, section)
    return bytes(pdf.output())


def _render_pdf_section(pdf, section: Section) -> None:
    from fpdf.enums import XPos, YPos

    if section.kind == "keyvalue":
        if section.title:
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(0, 7, _pdf_safe(section.title), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        for row in section.rows:
            label = row[0] if row else ""
            value = row[1] if len(row) > 1 else ""
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(75, 6, _pdf_safe(label))
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(0, 6, _pdf_safe(_fmt(value)), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(4)
        return

    if section.kind == "note":
        pdf.set_font("Helvetica", "I", 9)
        pdf.multi_cell(0, 5, _pdf_safe(section.note or (section.rows[0][0] if section.rows else "")))
        pdf.ln(2)
        return

    # table
    if not section.rows:
        return
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, _pdf_safe(section.title), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    ncol = max(len(r) for r in section.rows)
    usable = pdf.epw
    first_w = usable * (0.40 if ncol <= 2 else 0.28)
    rest_w = (usable - first_w) / (ncol - 1) if ncol > 1 else usable
    widths = [first_w] + [rest_w] * (ncol - 1)
    last_idx = len(section.rows) - 1
    for i, row in enumerate(section.rows):
        header = i == 0
        total = section.emphasize_last and i == last_idx
        pdf.set_font("Helvetica", "B" if header or total else "", 10)
        for c in range(ncol):
            cell = row[c] if c < len(row) else ""
            align = "L" if c == 0 else "R"
            is_last = c == ncol - 1
            pdf.cell(
                widths[c], 6, _pdf_safe(_fmt(cell)), border="B", align=align,
                new_x=XPos.LMARGIN if is_last else XPos.RIGHT,
                new_y=YPos.NEXT if is_last else YPos.TOP,
            )
    pdf.ln(3)


def render_xlsx(doc: Disclosure) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    sheet_map = doc.sheet_map or [(s.title or "Sheet", [s]) for s in doc.sections]
    first = True
    for sheet_name, sections in sheet_map:
        ws = wb.active if first else wb.create_sheet()
        ws.title = (sheet_name or "Sheet")[:31]
        first = False
        for section in sections:
            _render_xlsx_section(ws, section, Font)
            ws.append([])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _render_xlsx_section(ws, section: Section, Font) -> None:
    if section.title:
        ws.append([section.title])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    if section.kind == "note":
        ws.append([section.note or (section.rows[0][0] if section.rows else "")])
        return

    if section.kind == "keyvalue":
        for row in section.rows:
            ws.append(list(row))
        return

    # table
    last_idx = len(section.rows) - 1
    for i, row in enumerate(section.rows):
        ws.append(list(row))
        if i == 0 or (section.emphasize_last and i == last_idx):
            for col in range(1, len(row) + 1):
                ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
