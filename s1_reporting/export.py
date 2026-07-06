"""SB 253 / GHG Protocol disclosure exports.

Two artifacts over one locked inventory (research/2.4):
  - build_xlsx: structured, field-tagged workbook (the durable mapping layer for
    CARB's eventual 2027 template/portal).
  - build_pdf: a human-readable GHG-Protocol-aligned disclosure with a coversheet.
Both are pure: (InventoryReport, DisclosureMeta) -> bytes. fpdf2 is imported
lazily so the module loads without it.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO

from s1_reporting.report import InventoryReport


@dataclass
class DisclosureMeta:
    entity_name: str
    jurisdiction: str
    reporting_year: int
    period_start: str
    period_end: str
    consolidation_approach: str
    gwp_version: str
    base_year: int
    methodology_summary: str = (
        "Fuel-based (GHG Protocol Tier 1); EPA Emission Factors Hub 2025 / "
        "40 CFR Part 98; per-gas masses with GWP applied at reporting time."
    )


def _gas_rows(report: InventoryReport) -> list[tuple[str, float]]:
    g = report.by_gas
    return [
        ("CO2 (fossil)", g.co2_fossil_tco2e),
        ("CH4", g.ch4_tco2e),
        ("N2O", g.n2o_tco2e),
        ("SF6", g.sf6_tco2e),
        ("NF3", g.nf3_tco2e),
    ]


def _coversheet_rows(report: InventoryReport, meta: DisclosureMeta) -> list[tuple[str, object]]:
    return [
        ("Reporting entity", meta.entity_name),
        ("Jurisdiction", meta.jurisdiction),
        ("Reporting year", meta.reporting_year),
        ("Reporting period", f"{meta.period_start} to {meta.period_end}"),
        ("Consolidation approach", meta.consolidation_approach.replace("_", " ")),
        ("GWP version", meta.gwp_version),
        ("Base year", meta.base_year),
        ("Gross Scope 1 (tCO2e)", round(report.total_scope1_tco2e, 4)),
        ("Biogenic CO2 - memo, excluded (tCO2e)", round(report.biogenic_co2_tco2e, 4)),
        ("Records", report.record_count),
        ("Generated", date.today().isoformat()),
    ]


def build_xlsx(report: InventoryReport, meta: DisclosureMeta) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    cover = wb.active
    cover.title = "Disclosure"
    cover["A1"] = "CA SB 253 / GHG Protocol — Scope 1 Disclosure"
    cover["A1"].font = Font(bold=True, size=14)
    for label, value in _coversheet_rows(report, meta):
        cover.append([label, value])
    cover.append([])
    cover.append(["Methodology", meta.methodology_summary])

    gas = wb.create_sheet("By gas")
    gas.append(["Gas", "tCO2e"])
    for name, value in _gas_rows(report):
        gas.append([name, round(value, 6)])
    gas.append(["Total (gross Scope 1)", round(report.by_gas.total_tco2e, 6)])
    gas.append(["Biogenic CO2 (memo, excluded)", round(report.biogenic_co2_tco2e, 6)])

    fac = wb.create_sheet("By facility")
    fac.append(["Facility", "tCO2e"])
    for row in report.by_facility:
        fac.append([row.facility_name or "Unassigned", round(row.tco2e, 6)])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _pdf_safe(value: object) -> str:
    """fpdf2 core fonts are Latin-1 only; make arbitrary user text safe."""
    return str(value).encode("latin-1", "replace").decode("latin-1")


def build_pdf(report: InventoryReport, meta: DisclosureMeta) -> bytes:
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 9, "Scope 1 Emissions Disclosure", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(90, 90, 90)
    pdf.cell(0, 6, f"CA SB 253 / GHG Protocol Corporate Standard  ({report.ar_version})",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)

    for label, value in _coversheet_rows(report, meta):
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(75, 6, _pdf_safe(label))
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 6, _pdf_safe(value), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.ln(4)
    _pdf_table(pdf, "Emissions by gas", [("Gas", "tCO2e")]
               + [(n, f"{v:,.4f}") for n, v in _gas_rows(report)]
               + [("Total (gross Scope 1)", f"{report.by_gas.total_tco2e:,.4f}")])

    if report.by_facility:
        pdf.ln(3)
        _pdf_table(pdf, "Emissions by facility",
                   [("Facility", "tCO2e")]
                   + [(r.facility_name or "Unassigned", f"{r.tco2e:,.4f}") for r in report.by_facility])

    pdf.ln(4)
    pdf.set_font("Helvetica", "I", 9)
    pdf.multi_cell(0, 5, _pdf_safe(
        f"Biogenic CO2 of {report.biogenic_co2_tco2e:,.4f} tCO2e is reported as a "
        "separate memo line and is excluded from the gross Scope 1 total."))
    pdf.ln(1)
    pdf.multi_cell(0, 5, _pdf_safe(f"Methodology: {meta.methodology_summary}"))
    return bytes(pdf.output())


def _pdf_table(pdf, title: str, rows: list[tuple[str, str]]) -> None:
    from fpdf.enums import XPos, YPos

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, title, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    for i, (left, right) in enumerate(rows):
        header = i == 0
        pdf.set_font("Helvetica", "B" if header or left.startswith("Total") else "", 10)
        pdf.cell(130, 6, _pdf_safe(left), border="B")
        pdf.cell(0, 6, _pdf_safe(right), border="B", align="R", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
