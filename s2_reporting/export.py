"""Assurance-ready disclosure documents — XLSX + PDF (PRD V1 compliance).

Renders a ComplianceDisclosure (s2_reporting.compliance) into two artifacts an
assurer / filer actually uses:
  - build_disclosure_xlsx: a field-tagged workbook (Disclosure sheet = section /
    field / value / note; Readiness sheet = blockers + warnings) — the durable
    mapping layer.
  - build_disclosure_pdf: a human-readable filing with a readiness banner and one
    table per section.
Both are pure: (ComplianceDisclosure) -> bytes. openpyxl / fpdf2 are imported
lazily so the module loads without them. Imports only the compliance types.
"""

from __future__ import annotations

from io import BytesIO

from s2_reporting.compliance import ComplianceDisclosure


def _readiness_line(disclosure: ComplianceDisclosure) -> str:
    r = disclosure.readiness
    return "Assurance-ready — no blocking gaps" if r.ready else "NOT assurance-ready"


def build_disclosure_xlsx(disclosure: ComplianceDisclosure) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Disclosure"

    sheet["A1"] = f"{disclosure.standard_label} — Scope 2 Emissions Disclosure"
    sheet["A1"].font = Font(bold=True, size=14)
    for label, value in (
        ("Reporting entity", disclosure.entity),
        ("Reporting year", disclosure.reporting_year),
        ("Standard", disclosure.standard_label),
        ("Assurance readiness", _readiness_line(disclosure)),
    ):
        sheet.append([label, value])
    sheet.append([])

    header = sheet.max_row + 1
    sheet.append(["Section", "Field", "Value", "Note"])
    for cell in sheet[header]:
        cell.font = Font(bold=True)
    for section in disclosure.sections:
        for item in section.items:
            sheet.append([section.title, item.label, str(item.value), item.note or ""])

    ready = wb.create_sheet("Readiness")
    ready.append(["Assurance readiness", _readiness_line(disclosure)])
    ready.append([])
    ready.append(["Blockers"])
    for blocker in disclosure.readiness.blockers or ["(none)"]:
        ready.append(["", blocker])
    ready.append([])
    ready.append(["Warnings"])
    for warning in disclosure.readiness.warnings or ["(none)"]:
        ready.append(["", warning])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _pdf_safe(value: object) -> str:
    """fpdf2 core fonts are Latin-1 only; make arbitrary text safe."""
    return str(value).encode("latin-1", "replace").decode("latin-1")


def build_disclosure_pdf(disclosure: ComplianceDisclosure) -> bytes:
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 9, _pdf_safe(f"{disclosure.standard_label} — Scope 2 Disclosure"),
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(90, 90, 90)
    pdf.cell(0, 6, _pdf_safe(f"{disclosure.entity}  ·  reporting year {disclosure.reporting_year}"),
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    # Readiness banner.
    ready = disclosure.readiness
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(*(20, 120, 40) if ready.ready else (170, 40, 40))
    pdf.cell(0, 7, _pdf_safe(_readiness_line(disclosure)), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 9)
    for blocker in ready.blockers:
        pdf.multi_cell(0, 5, _pdf_safe(f"  - BLOCKER: {blocker}"),
                       new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    for warning in ready.warnings:
        pdf.multi_cell(0, 5, _pdf_safe(f"  - warning: {warning}"),
                       new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(2)

    for section in disclosure.sections:
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 7, _pdf_safe(section.title), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        for item in section.items:
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(95, 6, _pdf_safe(item.label), border="B")
            pdf.cell(0, 6, _pdf_safe(item.value), border="B", align="R",
                     new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            if item.note:
                pdf.set_font("Helvetica", "I", 8)
                pdf.set_text_color(140, 90, 20)
                pdf.multi_cell(0, 5, _pdf_safe(f"    {item.note}"),
                               new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                pdf.set_text_color(0, 0, 0)
        pdf.ln(2)

    return bytes(pdf.output())


__all__ = ["build_disclosure_xlsx", "build_disclosure_pdf"]
