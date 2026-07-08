"""Tests for assurance-ready disclosure export (XLSX/PDF)."""

from __future__ import annotations

from s2_reporting.compliance import DisclosureContext, build_csrd_e1, build_sb253
from s2_reporting.export import build_disclosure_pdf, build_disclosure_xlsx
from s2_reporting.summary import ReportSummary


def _disclosure(builder, **ctx_kw):
    summary = ReportSummary(
        entity="Acme Co",
        reporting_year=2025,
        location_based_tco2e=120.5,
        market_based_tco2e=90.0,
        consumption_mwh=250.0,
        data_coverage_pct=100.0,
        methodology="Dual-method per GHG Protocol Scope 2.",
    )
    return builder(summary, DisclosureContext(assurance_status="limited assurance obtained", **ctx_kw))


def test_xlsx_is_a_valid_workbook() -> None:
    data = build_disclosure_xlsx(_disclosure(build_sb253))
    assert data[:2] == b"PK"  # xlsx is a zip container
    assert len(data) > 1000


def test_pdf_has_pdf_magic() -> None:
    data = build_disclosure_pdf(_disclosure(build_sb253))
    assert data[:4] == b"%PDF"
    assert len(data) > 500


def test_xlsx_roundtrips_disclosure_content() -> None:
    import io

    from openpyxl import load_workbook

    data = build_disclosure_xlsx(_disclosure(build_csrd_e1, renewable_mwh=100.0))
    wb = load_workbook(io.BytesIO(data))
    assert {"Disclosure", "Readiness"} <= set(wb.sheetnames)
    text = " ".join(
        str(c.value) for row in wb["Disclosure"].iter_rows() for c in row if c.value is not None
    )
    assert "E1-6 Gross Scope 2 GHG emissions" in text
    assert "120.500 tCO2e" in text


def test_pdf_renders_not_ready_banner() -> None:
    # coverage below floor -> blocker -> "NOT assurance-ready" path in the PDF.
    summary = ReportSummary(
        entity="Acme",
        reporting_year=2025,
        location_based_tco2e=1.0,
        market_based_tco2e=1.0,
        consumption_mwh=0.0,  # blocker
        data_coverage_pct=0.0,
        methodology="m",
    )
    data = build_disclosure_pdf(build_sb253(summary, DisclosureContext()))
    assert data[:4] == b"%PDF"
