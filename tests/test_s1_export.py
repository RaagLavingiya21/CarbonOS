"""Tests for the SB 253 / GHG Protocol XLSX + PDF disclosure exports."""

from __future__ import annotations

from io import BytesIO

from fastapi.testclient import TestClient

from api.main import app
from s1_reporting import (
    DisclosureMeta,
    FacilityBreakdown,
    GasBreakdown,
    InventoryReport,
    build_pdf,
    build_xlsx,
)
from tests.conftest import AUTH_HEADERS

client = TestClient(app)


def _report() -> InventoryReport:
    return InventoryReport(
        ar_version="AR5",
        total_scope1_tco2e=5.31145,
        biogenic_co2_tco2e=0.5,
        by_gas=GasBreakdown(co2_fossil_tco2e=5.306, ch4_tco2e=0.0028, n2o_tco2e=0.00265),
        by_facility=[FacilityBreakdown("f1", "Café Plânt", 5.31145)],  # unicode -> must not crash PDF
        record_count=1,
    )


def _meta() -> DisclosureMeta:
    return DisclosureMeta(
        entity_name="Acme Mfg", jurisdiction="US", reporting_year=2025,
        period_start="2025-01-01", period_end="2025-12-31",
        consolidation_approach="operational_control", gwp_version="AR5", base_year=2025,
    )


# --- Pure builders ----------------------------------------------------------

def test_xlsx_is_a_valid_workbook() -> None:
    from openpyxl import load_workbook

    data = build_xlsx(_report(), _meta())
    assert data[:2] == b"PK"                       # xlsx = zip
    wb = load_workbook(BytesIO(data))
    assert set(wb.sheetnames) == {"Disclosure", "By gas", "By facility"}


def test_pdf_is_a_valid_document() -> None:
    data = build_pdf(_report(), _meta())
    assert data[:4] == b"%PDF"
    assert len(data) > 500


# --- Routes -----------------------------------------------------------------

def _mock_report_data(monkeypatch) -> None:
    monkeypatch.setattr("db.scope1_store.list_records_for_inventory", lambda inv, **k: [
        {"id": "r1", "inventory_id": inv, "emission_source_id": "s1",
         "kg_co2_fossil": 5306.0, "kg_ch4": 0.1, "kg_n2o": 0.01, "kg_co2_biogenic": 0.0}])
    monkeypatch.setattr("db.scope1_store.list_sources", lambda **k: [
        {"id": "s1", "entity_id": "e1", "facility_id": "f1", "source_name": "Boiler"}])
    monkeypatch.setattr("db.scope1_store.list_facilities", lambda **k: [{"id": "f1", "name": "Plant A"}])
    monkeypatch.setattr("db.scope1_store.list_boundaries", lambda inv, **k: [
        {"entity_id": "e1", "consolidation_multiplier": 1.0}])
    monkeypatch.setattr("db.scope1_store.get_inventory", lambda inv, **k: {
        "id": inv, "reporting_entity_id": "e1", "reporting_year": 2025,
        "period_start": "2025-01-01", "period_end": "2025-12-31",
        "consolidation_approach": "operational_control", "base_year": 2025})
    monkeypatch.setattr("db.scope1_store.get_entity", lambda eid, **k: {
        "id": eid, "name": "Acme Mfg", "jurisdiction": "US"})


def test_export_xlsx_route(monkeypatch) -> None:
    _mock_report_data(monkeypatch)
    resp = client.get("/api/scope1/inventories/inv1/report/xlsx", headers=AUTH_HEADERS)
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert resp.content[:2] == b"PK"
    assert ".xlsx" in resp.headers["content-disposition"]


def test_export_pdf_route(monkeypatch) -> None:
    _mock_report_data(monkeypatch)
    resp = client.get("/api/scope1/inventories/inv1/report/sb253.pdf", headers=AUTH_HEADERS)
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content[:4] == b"%PDF"
