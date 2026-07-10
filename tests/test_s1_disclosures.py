"""Pure tests for the multi-regime disclosure mappers + renderers (ESRS/CDP/GHGRP).

Rendering (fpdf2/openpyxl) + Latin-1 safety live in one place; these exercise the
mappers against a DisclosureData carrying non-zero fugitive + process so the
gross-Scope-1 correctness (combustion + fugitive + process) is verified.
"""

from __future__ import annotations

from io import BytesIO

from s1_calc.models import GasMasses
from s1_reporting import (
    DisclosureData,
    DisclosureMeta,
    FacilityBreakdown,
    GasBreakdown,
    InventoryReport,
    ReportRecord,
)
from s1_reporting.disclosures import map_cdp, map_esrs_e1, map_ghgrp, render_pdf, render_xlsx
from s1_reporting.disclosures.mappers import GHGRP_THRESHOLD_TCO2E, map_sb253
from s1_reporting.report import SourceLine


def _data(combustion_total: float = 100.0, fugitive: float = 12.0, process: float = 8.0) -> DisclosureData:
    rep = InventoryReport(
        ar_version="AR5", total_scope1_tco2e=combustion_total, biogenic_co2_tco2e=5.0,
        by_gas=GasBreakdown(co2_fossil_tco2e=90.0, ch4_tco2e=6.0, n2o_tco2e=4.0),
        by_facility=[FacilityBreakdown("f1", "Café Plânt", combustion_total)], record_count=1,
    )
    recs = [ReportRecord(record_id="r1", gas_masses=GasMasses(kg_co2_fossil=90000, kg_ch4=200, kg_n2o=1),
                         facility_name="Café Plânt", fuel="natural_gas", ef_tier="T2")]
    return DisclosureData(
        combustion=rep, fugitive_tco2e=fugitive, process_tco2e=process,
        fugitive_lines=[SourceLine("R-410A", "R-410A", "f1", "Café Plânt", fugitive)],
        process_lines=[SourceLine("cement_clinker", "Carbon dioxide", "f1", "Café Plânt", process)],
        report_records=recs,
    )


def _meta() -> DisclosureMeta:
    return DisclosureMeta(
        entity_name="Acme", jurisdiction="US", reporting_year=2024,
        period_start="2024-01-01", period_end="2024-12-31",
        consolidation_approach="operational_control", gwp_version="AR5", base_year=2020,
    )


def _rows(section):
    return section.rows


def test_gross_includes_all_three_categories() -> None:
    d = _data()
    assert d.gross_scope1_tco2e == 120.0        # 100 + 12 + 8
    # every regime's gross figure reflects it
    for mapper in (map_sb253, map_esrs_e1, map_cdp):
        doc = mapper(d, _meta())
        flat = [str(c) for s in doc.sections for r in s.rows for c in r]
        assert any("120" in v for v in flat), doc.regime


def test_esrs_has_e16_boundary_biogenic_methodology() -> None:
    doc = map_esrs_e1(_data(), _meta())
    titles = " ".join(s.title for s in doc.sections)
    assert "E1-6" in titles
    notes = " ".join(s.note for s in doc.sections if s.kind == "note")
    assert "Biogenic" in notes and "Methodology" in notes
    assert any(s.kind == "table" and "gas" in s.title.lower() for s in doc.sections)


def test_assurance_line_defaults_to_placeholder() -> None:
    from s1_reporting.disclosures.mappers import assurance_line
    assert assurance_line(_meta()) == "Not yet third-party verified"     # no assurance set


def test_assurance_line_reports_real_status() -> None:
    from s1_reporting.disclosures.mappers import assurance_line
    m = DisclosureMeta(
        entity_name="Acme", jurisdiction="US", reporting_year=2024,
        period_start="2024-01-01", period_end="2024-12-31",
        consolidation_approach="operational_control", gwp_version="AR5", base_year=2020,
        assurance_level="limited", assurance_standard="ISAE_3410",
        assurance_statement_on_file=True,
    )
    line = assurance_line(m)
    assert line == "Limited assurance (ISAE 3410); assurance statement on file"
    # and it flows into every regime's verification section
    for mapper in (map_sb253, map_esrs_e1, map_cdp, map_ghgrp):
        doc = mapper(_data(), m)
        assert any(line in str(c) for s in doc.sections for r in (s.rows or []) for c in r)


def test_cdp_has_bygas_byfacility_verification() -> None:
    doc = map_cdp(_data(), _meta())
    titles = " ".join(s.title for s in doc.sections).lower()
    assert "c6.1" in titles and "by facility" in titles and "verification" in titles
    verif = next(s for s in doc.sections if "verification" in s.title.lower())
    assert any("Not yet third-party verified" in str(c) for r in verif.rows for c in r)


def test_ghgrp_units_table_and_threshold_flag() -> None:
    # over threshold
    over = map_ghgrp(_data(combustion_total=30_000.0), _meta())
    note = " ".join(s.note for s in over.sections if s.kind == "note")
    assert "AT OR ABOVE" in note
    units = next(s for s in over.sections if "combustion units" in s.title.lower())
    assert units.rows[0] == ["Facility", "Fuel", "CO2 (kg)", "CH4 (kg)", "N2O (kg)", "Tier"]
    data_row = units.rows[1]
    assert data_row[0] == "Café Plânt" and data_row[1] == "natural_gas"
    assert data_row[2] == 90000.0 and data_row[5] == "T2"   # CO2 kg + tier

    # under threshold
    under = map_ghgrp(_data(combustion_total=1.0, fugitive=0.0, process=0.0), _meta())
    assert "below" in " ".join(s.note for s in under.sections if s.kind == "note")


def test_threshold_constant() -> None:
    assert GHGRP_THRESHOLD_TCO2E == 25_000.0


def test_renderers_produce_valid_bytes_with_unicode() -> None:
    """Café Plânt (Unicode) must not crash the Latin-1 PDF; xlsx is a valid zip."""
    for mapper in (map_sb253, map_esrs_e1, map_cdp, map_ghgrp):
        doc = mapper(_data(), _meta())
        pdf = render_pdf(doc)
        assert pdf[:4] == b"%PDF" and len(pdf) > 400
        xlsx = render_xlsx(doc)
        assert xlsx[:2] == b"PK"
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(xlsx))
        assert wb.sheetnames                      # at least one sheet
